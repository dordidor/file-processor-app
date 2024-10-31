import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from io import BytesIO
import yfinance as yf

def detect_highlighted_cells(filename, sheet_name, highlight_color='#ffff00'):
    # Load the workbook and select the worksheet
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheet_name]
    
    # Dictionary to store highlighted cells and their values
    highlighted_cells = {}

    # Iterate over all cells in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            # Check if the cell has a fill color
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                # Compare the cell's color to the highlight color
                cell_color = cell.fill.start_color.rgb[-6:]  # Get RGB color without alpha
                if cell_color == highlight_color:
                    highlighted_cells[cell.coordinate] = cell.value  # Store cell address and value

    return highlighted_cells

def highlight_winning_bids(cell, winning_dict, block_height, miner_id):
    """Apply highlight color to cells if they match the winning bid."""
    if block_height in winning_dict:
        winner_miner = winning_dict[block_height]['winning bidder']
        winning_bid = winning_dict[block_height]['winning bid']
        if miner_id == winner_miner and cell == winning_bid:
            return 'background-color: yellow'
    return ''

def process_data(df):
    # Drop rows with missing 'block height'
    df = df.dropna(subset=['block height'])
    df = df.loc[:, ~df.columns.str.contains('^Unnamed:')]

    # Melt the dataframe to get MinerID and corresponding Bid
    miner_ids = pd.melt(df, id_vars='block height', value_vars=['miner 1','miner 2','miner 3','miner 4','miner 5','miner 6'],  value_name='MinerID')
    miner_ids['Bid'] = pd.melt(df, id_vars='block height', value_vars=['bid 1','bid 2','bid 3','bid 4','bid 5','bid 6'], value_name='Bid')['Bid']
    
    # Calculate the total bid per block height
    miner_ids['Total'] = miner_ids.groupby('block height')['Bid'].transform('sum')
    miner_ids = miner_ids.dropna(subset=['MinerID', 'Bid'], how='all')

    # Create a dictionary of winning bids using the 'winning bidder' column
    auction_winners = dict(zip(df['block height'], df['winning bidder']))

    return miner_ids

def pivot_data(df):
    # Pivot the data to get miner IDs as columns
    df_wide = df.pivot(index=['block height','Total'], columns='MinerID', values='Bid').reset_index()
    
    return df_wide

# Step 2: Define a function to read the files
def read_file(file):
    if file is not None:
        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        elif file.name.endswith('.xlsx'):
            df = pd.read_excel(file)
        return df
    return None

def read_data_with_highlights(filename, sheet_name, highlight_color_hex='FFFF00'):
    # Load the workbook and select the worksheet
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheet_name]
    
    # Detect the header row (first row) and initialize an empty list to store data
    header = [cell.value for cell in ws[1]]  # Assumes the first row is the header
    auction_data = []

    # Iterate over each row after the header
    for row in ws.iter_rows(min_row=2):
        auction_id = row[0].value  # First cell in each row is the auction ID
        row_data = {"block height": auction_id}

        # Loop over each bidder column and detect the bid amount and highlight status
        for cell in row[1:]:  # Skip the first cell (auction_id)
            bidder_id = header[cell.column - 1]  # Get the bidder ID from the header
            bid_amount = cell.value
            
            # Check if the cell has a fill color matching the highlight color
            is_winner = False
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                cell_color_hex = cell.fill.start_color.rgb[-6:]  # Extract RGB hex without alpha
                if cell_color_hex.upper() == highlight_color_hex.upper():
                    is_winner = True
            
            # Store data in the row dictionary with is_winner flag
            row_data["MinerID"] = bidder_id
            row_data["Bid"] = bid_amount
            row_data["is_winner"] = is_winner
            
            # Append row data to auction data
            auction_data.append(row_data.copy())

    # Convert the list of dictionaries to a DataFrame
    df_long = pd.DataFrame(auction_data)
    df_long = df_long[df_long["MinerID"] != 'Total']

    return df_long

# Step 1: Take user input for two files
st.title("File Processor and Appender")

tab_titles = ['Load New Data', 'Statistics', 'Graphs']
tabs = st.tabs(tab_titles)

with tabs[0]:

    st.markdown("#### This tab processes new data and appends to master sheet")

    # File uploader for the file (file to process)
    file1 = st.file_uploader("Upload the first file (file to process)", type=['csv', 'xlsx'])
    file2 = st.file_uploader("Upload the second file (file to append to)", type=['csv', 'xlsx'])

    # Step 3: Process the files
    if file1 and file2:
        df1 = read_file(file1)
        df2 = read_file(file2)

        if df1 is not None and df2 is not None:
            st.write("File 1 (to process):")
            st.dataframe(df1.head())  

            st.write("File 2 (to append to):")
            st.dataframe(df2.head())  

            # Process file 1 and extract winning bids
            processed_df1 = process_data(df1)  
            df_wide = pivot_data(processed_df1)

            # Append processed data to file 2
            appended_df = pd.concat([df2, df_wide]).reset_index(drop=True)

            # Display the appended dataframe
            st.write("Appended DataFrame:")
            st.dataframe(appended_df)

            # # Step 4: Apply conditional formatting for display in Streamlit
            # def apply_highlight(value, row, col):
            #     block_height = row['block height']
            #     return highlight_winning_bids(value, winning_dict, block_height, col)

            # # Create a function to highlight winning bids in the DataFrame
            # def highlight_row(row):
            #     styles = []
            #     for col in appended_df.columns:
            #         if col not in ['block height', 'Total']:  # Apply only to miner bid columns
            #             style = apply_highlight(row[col], row, col) if col in row else ''
            #             styles.append(style)
            #         else:
            #             styles.append('')  # Empty style for non-bid columns
            #     return styles
            
            # # Apply row-wise styling
            # highlighted_df = appended_df.style.apply(highlight_row, axis=1)

            # # Display the appended dataframe
            # st.write("Appended DataFrame:")
            # st.write(highlighted_df)

            # Save the data to an Excel file
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                appended_df.to_excel(writer, index=False, sheet_name='AppendedData')
                worksheet = writer.sheets['AppendedData']
            #     highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            #     for row in range(2, worksheet.max_row + 1):  # Skip header row
            #         block_height = worksheet.cell(row=row, column=1).value
            #         if block_height in winning_dict:
            #             winner_miner = winning_dict[block_height]['winning bidder']
            #             winning_bid = winning_dict[block_height]['winning bid']
            #             for col in range(3, worksheet.max_column + 1):  # Start from column with bids
            #                 miner_id = worksheet.cell(1, col).value
            #                 if miner_id == winner_miner:
            #                     cell = worksheet.cell(row=row, column=col)
            #                     if cell.value == winning_bid:
            #                         cell.fill = highlight_fill

            # Convert in-memory file to downloadable format
            output.seek(0)
            st.download_button(
                label="Download appended file",
                data=output,
                file_name='appended_file.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            st.error("One of the files could not be read. Please check the format.")

with tabs[1]:

    st.markdown("#### This tab does some analysis of the imported data")

    # File uploader for the file (file to process)
    file1 = st.file_uploader("Analyze this file:", type=['csv', 'xlsx'])

    # Check if a file is uploaded
    if file1 is not None:
        # If the uploaded file is an Excel file
        if file1.name.endswith('.xlsx'):
            # Load the workbook to get the sheet names
            xls = pd.ExcelFile(file1)
            sheet_names = xls.sheet_names
            
            # Dropdown to select the sheet
            selected_sheet = st.selectbox("Select a worksheet", sheet_names)
            
            # Read the selected sheet into a DataFrame
            df_long = read_data_with_highlights(file1, selected_sheet)
        else:
            # If it's a CSV, read directly
            # df = pd.read_csv(file1)
            print("File is not an excel file")
        
        summary = df_long.groupby('MinerID').agg(
                                                    Total_Bid_BTC = ('Bid', 'sum'),
                                                    Avg_Bid_BTC=('Bid', 'mean')
                                                ).reset_index()

        summary.rename(columns={'Total_Bid_BTC': 'Total Bid (BTC)', 'Avg_Bid_BTC': 'Average Bid (BTC)'}, inplace=True)

        summary['Total Bid (BTC)'] =  2 * summary['Total Bid (BTC)']
        summary['Average Bid (BTC)'] =  2 * summary['Average Bid (BTC)']

        # Number input
        btc = st.number_input("Price of BTC:", min_value=0)
        satoshi = st.number_input("Price of Satoshi:", min_value=0)

        summary['Total Bid (USD)'] = summary['Total Bid (BTC)']/satoshi * btc
        summary['Average Bid (USD)'] = summary['Average Bid (BTC)']/satoshi * btc

        st.dataframe(summary)

with tabs[2]:

    # Bitcoin ticker on Yahoo Finance is BTC-USD
    ticker = 'BTC-USD'
    # data = yf.download(ticker, period='1d', interval='1m')

    # Get the latest price (most recent close price)
    # latest_price = data['Close'][-1]
    # print(f"Current Bitcoin price: ${latest_price}")
